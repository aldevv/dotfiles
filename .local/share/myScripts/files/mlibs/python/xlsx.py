from openpyxl import load_workbook


def XLSXDictReader(f):
    book = load_workbook(f)
    sheet = book.active
    rows = sheet.max_row
    cols = sheet.max_column
    headers = dict((i, sheet.cell(row=1, column=i).value) for i in range(1, cols))

    def item(i, j):
        return (sheet.cell(row=1, column=j).value, sheet.cell(row=i, column=j).value)

    return (dict(item(i, j) for j in range(1, cols + 1)) for i in range(2, rows + 1))


def XLSXListReader(f):
    book = load_workbook(f)
    sheet = book.active
    rows = sheet.max_row
    cols = sheet.max_column
    headers = dict((i, sheet.cell(row=1, column=i).value) for i in range(1, cols))

    def item(i, j):
        return (sheet.cell(row=1, column=j).value, sheet.cell(row=i, column=j).value)

    return (list(item(i, j)[1] for j in range(1, cols + 1)) for i in range(2, rows + 1))


def getHeaders(f):
    book = load_workbook(f)
    sheet = book.active
    rows = sheet.max_row
    cols = sheet.max_column
    headers = dict((i, sheet.cell(row=1, column=i).value) for i in range(1, cols))

    return headers
